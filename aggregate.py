import openpyxl
import json

data = {}

def getCellValue(row, col, ws):
        cell = ws.cell(row, col)
        val = cell.value
        return val

def scanExcelOld(year):
    try:
        wb = openpyxl.load_workbook(f'State_to_State_Migrations_Table_{year}.xlsx')
    except FileNotFoundError as e:
        print(Exception(f"Invalid file name was entered with year: {year}"))
        return
    ws = wb['Table']

    cur_row = 10
    cur_col = 2

    for i in range(0,11):
        for j in range(0,5):
            if cur_row == 19: # District of Columbia
                cur_row += 1
                continue
            if cur_row == 60: # Puerto Rico
                cur_row += 1
                continue
            if cur_row == 76: # no more data
                break
            if cur_row == 46: # break in data
                cur_row = 50
            
            stateName = getCellValue(cur_row, 1, ws)

            fromStateData = {}

            for k in range(0,11):
                for l in range(0,5):
                    if cur_col == 19: # District of Columbia
                        cur_col += 2
                        continue
                    if cur_col == 87: # Puerto Rico
                        cur_col += 2
                        continue
                    if cur_col == 141: # end of data
                        break

                    singleData = {}

                    singleData['estimate'] = getCellValue(cur_row, cur_col, ws)
                    cur_col += 1
                    singleData['MOE'] = getCellValue(cur_row, cur_col, ws)
                    cur_col += 1

                    fromStateData[getCellValue(7, cur_col - 2,ws)] = singleData
                cur_col += 1
            cur_col = 2
            if year == 2005:
                data[stateName] = {}
            data[stateName][year] = fromStateData

            cur_row += 1
        cur_row += 1
    return

def scanExcelNew(year):
    try:
        wb = openpyxl.load_workbook(f'State_to_State_Migrations_Table_{year}.xlsx')
    except FileNotFoundError as e:
        print(Exception(f"Invalid file name was entered with year: {year}"))
        return
    ws = wb['Table']

    cur_row = 12
    cur_col = 10

    for i in range(0,11):
        for j in range(0,5):
            if cur_row == 21: # District of Columbia
                cur_row += 1
                continue
            if cur_row == 44: # Break in table
                cur_row = 48
                break
            if cur_row == 77: # No more data
                break
            
            stateName = getCellValue(cur_row, 1, ws)

            fromStateData = {}

            singleData = {}
            singleData['estimate'] = getCellValue(cur_row, cur_col,ws)
            cur_col += 1
            singleData['MOE'] = getCellValue(cur_row, cur_col,ws)
            cur_col += 2

            fromStateData[getCellValue(7, cur_col - 3, ws)] = singleData # {'State': {'estimate': 'N/A', 'MOE': 'N/A'}

            for k in range(0,10):
                for l in range(0,5):
                    if cur_col == 28: # District of Columbia
                        cur_col += 2
                        continue

                    singleData = {}

                    singleData['estimate'] = getCellValue(cur_row, cur_col,ws)
                    cur_col += 1
                    singleData['MOE'] = getCellValue(cur_row, cur_col,ws)
                    cur_col += 1

                    fromStateData[getCellValue(7, cur_col - 2,ws)] = singleData
                    
                cur_col += 1
            cur_col = 10
            data[stateName][year] = fromStateData

            cur_row += 1
        cur_row += 1
    return 

for i in range(2005,2010):
    scanExcelOld(i)

for i in range(2010,2023):
    scanExcelNew(i)

with open ('data.json', 'w') as output:
    json.dump(data, output)